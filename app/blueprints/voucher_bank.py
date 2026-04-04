# -*- coding: utf-8 -*-
"""
証憑データ化アプリ - 通帳モード Blueprint
通帳画像のアップロード、OCR処理（バックグラウンド）、一覧表示、CSV/Excel出力
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import os
import io
import csv
import threading
from datetime import datetime
from app.db import SessionLocal
from app.models_voucher import TBankStatement, TBankTransaction, TBankColumnTemplate, TBankDescriptionLearning
from app.models_login import TTenpo, TTenant, TKanrisha, TAppManagerGroup
from app.utils.decorators import require_roles, ROLES
from app.utils.voucher.ocr import process_bank_statement_image, save_uploaded_file

bp = Blueprint('voucher_bank', __name__, url_prefix='/voucher/bank')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _normalize_description(text: str) -> str:
    """摘要文字列を正規化（全角→半角、大文字→小文字、スペース除去）して照合精度を上げる"""
    if not text:
        return ''
    import unicodedata
    # NFKC正規化（全角→半角、ユニコード正規化）
    text = unicodedata.normalize('NFKC', text)
    # 大文字→小文字
    text = text.lower()
    # 先頭・末尾・連続スペースを除去
    text = ' '.join(text.split())
    return text


def _save_learning_data(db, tenant_id: int, original_desc: str, corrected_desc: str, tenpo_id: int = None, template_id: int = None):
    """手動修正の摘要対応を学習テーブルに保存または更新する（店舗・テンプレート単位で分離）"""
    if not original_desc or not corrected_desc:
        return
    if original_desc == corrected_desc:
        return  # 変更なしは記録しない
    norm_orig = _normalize_description(original_desc)
    if not norm_orig:
        return
    # 既存レコードを検索（店舗・テンプレート・元摘要・修正摘要の組み合わせ）
    q = db.query(TBankDescriptionLearning).filter(
        TBankDescriptionLearning.tenant_id == tenant_id,
        TBankDescriptionLearning.元摘要 == norm_orig,
        TBankDescriptionLearning.修正摘要 == corrected_desc
    )
    if tenpo_id is not None:
        q = q.filter(TBankDescriptionLearning.tenpo_id == tenpo_id)
    else:
        q = q.filter(TBankDescriptionLearning.tenpo_id.is_(None))
    if template_id is not None:
        q = q.filter(TBankDescriptionLearning.template_id == template_id)
    else:
        q = q.filter(TBankDescriptionLearning.template_id.is_(None))
    existing = q.first()
    if existing:
        existing.適用回数 += 1
        print(f'[学習] 適用回数更新: "{original_desc}" → "{corrected_desc}" (店舗={tenpo_id}, テンプレート={template_id}, 回数={existing.適用回数})')
    else:
        new_record = TBankDescriptionLearning(
            tenant_id=tenant_id,
            tenpo_id=tenpo_id,
            template_id=template_id,
            元摘要=norm_orig,
            修正摘要=corrected_desc,
            適用回数=1
        )
        db.add(new_record)
        print(f'[学習] 新規学習: "{original_desc}" → "{corrected_desc}" (店舗={tenpo_id}, テンプレート={template_id})')


def get_session_info():
    tenpo_id = session.get('store_id') or session.get('tenpo_id')
    return {
        'tenant_id': session.get('tenant_id'),
        'tenpo_id': tenpo_id,
        'user_id': session.get('user_id'),
        'role': session.get('role'),
    }


def get_openai_api_key(tenant_id, tenpo_id):
    """店舗 → テナント → システム管理者の順でOpenAI APIキーを取得"""
    keys = get_api_keys(tenant_id, tenpo_id)
    return keys.get('openai_api_key')


def get_api_keys(tenant_id, tenpo_id):
    """店舗 → テナント → システム管理者の順で各種 APIキーを取得"""
    db = SessionLocal()
    result = {'openai_api_key': None, 'google_vision_api_key': None, 'google_api_key': None, 'anthropic_api_key': None,
              'azure_document_intelligence_endpoint': None, 'azure_document_intelligence_key': None}
    try:
        if tenpo_id:
            tenpo = db.query(TTenpo).filter(TTenpo.id == tenpo_id).first()
            if tenpo:
                if getattr(tenpo, 'openai_api_key', None):
                    result['openai_api_key'] = tenpo.openai_api_key
                if getattr(tenpo, 'google_vision_api_key', None):
                    result['google_vision_api_key'] = tenpo.google_vision_api_key
                if getattr(tenpo, 'google_api_key', None):
                    result['google_api_key'] = tenpo.google_api_key
                if getattr(tenpo, 'anthropic_api_key', None):
                    result['anthropic_api_key'] = tenpo.anthropic_api_key
                if getattr(tenpo, 'azure_document_intelligence_endpoint', None):
                    result['azure_document_intelligence_endpoint'] = tenpo.azure_document_intelligence_endpoint
                if getattr(tenpo, 'azure_document_intelligence_key', None):
                    result['azure_document_intelligence_key'] = tenpo.azure_document_intelligence_key
        if tenant_id:
            tenant = db.query(TTenant).filter(TTenant.id == tenant_id).first()
            if tenant:
                if not result['openai_api_key'] and getattr(tenant, 'openai_api_key', None):
                    result['openai_api_key'] = tenant.openai_api_key
                if not result['google_vision_api_key'] and getattr(tenant, 'google_vision_api_key', None):
                    result['google_vision_api_key'] = tenant.google_vision_api_key
                if not result['google_api_key'] and getattr(tenant, 'google_api_key', None):
                    result['google_api_key'] = tenant.google_api_key
                if not result['anthropic_api_key'] and getattr(tenant, 'anthropic_api_key', None):
                    result['anthropic_api_key'] = tenant.anthropic_api_key
                if not result['azure_document_intelligence_endpoint'] and getattr(tenant, 'azure_document_intelligence_endpoint', None):
                    result['azure_document_intelligence_endpoint'] = tenant.azure_document_intelligence_endpoint
                if not result['azure_document_intelligence_key'] and getattr(tenant, 'azure_document_intelligence_key', None):
                    result['azure_document_intelligence_key'] = tenant.azure_document_intelligence_key
        # アプリ管理グループから未設定のAPIキーをフォールバック
        needs_group_fallback = not all([result['openai_api_key'], result['google_vision_api_key'],
                                        result['azure_document_intelligence_endpoint'], result['azure_document_intelligence_key']])
        if needs_group_fallback and tenant_id:
            # テナントに結びついたアプリ管理者を取得してグループを確認
            try:
                app_managers = db.query(TKanrisha).filter(
                    TKanrisha.role == 'app_manager',
                    TKanrisha.app_manager_group_id.isnot(None)
                ).all()
                for am in app_managers:
                    if am.app_manager_group_id:
                        grp = db.query(TAppManagerGroup).filter(TAppManagerGroup.id == am.app_manager_group_id).first()
                        if grp:
                            if not result['openai_api_key'] and getattr(grp, 'openai_api_key', None):
                                result['openai_api_key'] = grp.openai_api_key
                            if not result['google_vision_api_key'] and getattr(grp, 'google_vision_api_key', None):
                                result['google_vision_api_key'] = grp.google_vision_api_key
                            if not result['google_api_key'] and getattr(grp, 'google_api_key', None):
                                result['google_api_key'] = grp.google_api_key
                            if not result['anthropic_api_key'] and getattr(grp, 'anthropic_api_key', None):
                                result['anthropic_api_key'] = grp.anthropic_api_key
                            if not result['azure_document_intelligence_endpoint'] and getattr(grp, 'azure_document_intelligence_endpoint', None):
                                result['azure_document_intelligence_endpoint'] = grp.azure_document_intelligence_endpoint
                            if not result['azure_document_intelligence_key'] and getattr(grp, 'azure_document_intelligence_key', None):
                                result['azure_document_intelligence_key'] = grp.azure_document_intelligence_key
                            break
            except Exception:
                pass
        # システム管理者から未設定のAPIキーをフォールバック
        needs_fallback = not all([result['openai_api_key'], result['google_vision_api_key'], result['google_api_key'], result['anthropic_api_key']])
        if needs_fallback:
            sys_admins = db.query(TKanrisha).filter(
                TKanrisha.role == 'system_admin'
            ).all()
            for sys_admin in sys_admins:
                if not result['openai_api_key'] and getattr(sys_admin, 'openai_api_key', None):
                    result['openai_api_key'] = sys_admin.openai_api_key
                if not result['google_vision_api_key'] and getattr(sys_admin, 'google_vision_api_key', None):
                    result['google_vision_api_key'] = sys_admin.google_vision_api_key
                if not result['google_api_key'] and getattr(sys_admin, 'google_api_key', None):
                    result['google_api_key'] = sys_admin.google_api_key
                if not result['anthropic_api_key'] and getattr(sys_admin, 'anthropic_api_key', None):
                    result['anthropic_api_key'] = sys_admin.anthropic_api_key
                if not result['azure_document_intelligence_endpoint'] and getattr(sys_admin, 'azure_document_intelligence_endpoint', None):
                    result['azure_document_intelligence_endpoint'] = sys_admin.azure_document_intelligence_endpoint
                if not result['azure_document_intelligence_key'] and getattr(sys_admin, 'azure_document_intelligence_key', None):
                    result['azure_document_intelligence_key'] = sys_admin.azure_document_intelligence_key
                if all([result['openai_api_key'], result['google_vision_api_key'], result['google_api_key'], result['anthropic_api_key']]):
                    break
    except Exception:
        pass
    finally:
        db.close()
    return result


def _run_ocr_background(stmt_id, filepath, api_key, tenant_id, google_vision_api_key=None, column_def=None, tenpo_id=None, template_id=None, azure_document_intelligence_endpoint=None, azure_document_intelligence_key=None):
    """バックグラウンドスレッドでOCR処理を実行してDBを更新する"""
    db = SessionLocal()
    try:
        stmt = db.query(TBankStatement).filter(TBankStatement.id == stmt_id).first()
        if not stmt:
            return
        stmt.ステータス = 'processing'
        db.commit()

        ocr_result = process_bank_statement_image(filepath, api_key=api_key, google_vision_api_key=google_vision_api_key, column_def=column_def,
                                                    azure_document_intelligence_endpoint=azure_document_intelligence_endpoint,
                                                    azure_document_intelligence_key=azure_document_intelligence_key)

        stmt = db.query(TBankStatement).filter(TBankStatement.id == stmt_id).first()
        if not stmt:
            return

        stmt.OCR結果_生データ = ocr_result.get('raw_text', '')
        stmt.銀行名 = ocr_result.get('bank_name')
        stmt.支店名 = ocr_result.get('branch_name')
        stmt.口座種別 = ocr_result.get('account_type')
        stmt.口座番号 = ocr_result.get('account_number')
        stmt.口座名義 = ocr_result.get('account_holder')
        stmt.期間_開始 = ocr_result.get('period_start')
        stmt.期間_終了 = ocr_result.get('period_end')
        stmt.ステータス = 'completed'
        db.flush()

        db.query(TBankTransaction).filter(TBankTransaction.statement_id == stmt_id).delete()

        # 学習データを取得して自動適用用のマッピングを作成
        # 店舗・テンプレート単位で学習データを絞り込む
        learning_map = {}
        try:
            lq = db.query(TBankDescriptionLearning).filter(
                TBankDescriptionLearning.tenant_id == tenant_id
            )
            if tenpo_id is not None:
                lq = lq.filter(TBankDescriptionLearning.tenpo_id == tenpo_id)
            else:
                lq = lq.filter(TBankDescriptionLearning.tenpo_id.is_(None))
            if template_id is not None:
                lq = lq.filter(TBankDescriptionLearning.template_id == template_id)
            else:
                lq = lq.filter(TBankDescriptionLearning.template_id.is_(None))
            learnings = lq.all()
            for lrn in learnings:
                # 元摘要を正規化（全角・半角・大小文字統一）してマッピング
                key = _normalize_description(lrn.元摘要)
                # 同じ元摘要に複数の修正がある場合は適用回数が多いものを優先
                if key not in learning_map or lrn.適用回数 > learning_map[key].適用回数:
                    learning_map[key] = lrn
        except Exception as e:
            print(f'[学習] 学習データ取得エラー: {e}')
            # DBセッションがエラー状態になっている可能性があるためrollbackして
            # completedステータスを確実に保存する
            try:
                db.rollback()
                stmt2 = db.query(TBankStatement).filter(TBankStatement.id == stmt_id).first()
                if stmt2:
                    stmt2.OCR結果_生データ = ocr_result.get('raw_text', '')
                    stmt2.銀行名 = ocr_result.get('bank_name')
                    stmt2.支店名 = ocr_result.get('branch_name')
                    stmt2.口座種別 = ocr_result.get('account_type')
                    stmt2.口座番号 = ocr_result.get('account_number')
                    stmt2.口座名義 = ocr_result.get('account_holder')
                    stmt2.期間_開始 = ocr_result.get('period_start')
                    stmt2.期間_終了 = ocr_result.get('period_end')
                    stmt2.ステータス = 'completed'
                    db.flush()
                    db.query(TBankTransaction).filter(TBankTransaction.statement_id == stmt_id).delete()
                    for i, t in enumerate(ocr_result.get('transactions', []), 1):
                        row = TBankTransaction(
                            statement_id=stmt_id,
                            tenant_id=tenant_id,
                            日付=t.get('date'),
                            摘要=t.get('description'),
                            手書き摘要=t.get('note'),
                            入金=t.get('deposit'),
                            出金=t.get('withdrawal'),
                            残高=t.get('balance'),
                            行番号=i,
                        )
                        db.add(row)
                    db.commit()
                    print(f'[学習] 学習データなしでOCR結果を保存しました（{len(ocr_result.get("transactions", []))}件）')
            except Exception as e2:
                print(f'[学習] フォールバック保存エラー: {e2}')
            return

        applied_count = 0
        for i, t in enumerate(ocr_result.get('transactions', []), 1):
            desc = t.get('description') or ''
            # 学習データと照合して自動適用
            norm_desc = _normalize_description(desc)
            if norm_desc and norm_desc in learning_map:
                corrected = learning_map[norm_desc].修正摘要
                print(f'[学習] 摘要自動適用: "{desc}" → "{corrected}"')
                t['description'] = corrected
                applied_count += 1
            row = TBankTransaction(
                statement_id=stmt_id,
                tenant_id=tenant_id,
                日付=t.get('date'),
                摘要=t.get('description'),
                手書き摘要=t.get('note'),
                入金=t.get('deposit'),
                出金=t.get('withdrawal'),
                残高=t.get('balance'),
                行番号=i,
            )
            db.add(row)
        if applied_count > 0:
            print(f'[学習] 合計{applied_count}行の摘要を自動適用しました')

        db.commit()

    except Exception as e:
        try:
            stmt = db.query(TBankStatement).filter(TBankStatement.id == stmt_id).first()
            if stmt:
                stmt.ステータス = 'error'
                stmt.OCR結果_生データ = f'OCRエラー: {str(e)}'
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


# ============================================================
# 通帳明細一覧
# ============================================================
@bp.route('/')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def index():
    info = get_session_info()
    tenant_id = info['tenant_id']
    tenpo_id = info['tenpo_id']
    role = info['role']

    if not tenant_id:
        flash('テナントが選択されていません', 'error')
        return redirect(url_for('auth.select_login'))

    db = SessionLocal()
    try:
        q = db.query(TBankStatement).filter(TBankStatement.tenant_id == tenant_id)
        if role in (ROLES['ADMIN'], ROLES['EMPLOYEE']) and tenpo_id:
            q = q.filter(TBankStatement.tenpo_id == tenpo_id)
        statements = q.order_by(TBankStatement.created_at.desc()).all()
    finally:
        db.close()

    return render_template('voucher_bank_list.html', statements=statements)


# ============================================================
# 通帳アップロード（即時リダイレクト、OCRはバックグラウンド）
# ============================================================
@bp.route('/upload', methods=['GET', 'POST'])
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def upload():
    if request.method == 'GET':
        info = get_session_info()
        tenant_id = info['tenant_id']
        templates = []
        if tenant_id:
            db = SessionLocal()
            try:
                templates = db.query(TBankColumnTemplate).filter(
                    TBankColumnTemplate.tenant_id == tenant_id
                ).order_by(TBankColumnTemplate.is_default.desc(), TBankColumnTemplate.id.desc()).all()
            finally:
                db.close()
        return render_template('voucher_bank_upload.html', templates=templates)

    info = get_session_info()
    tenant_id = info['tenant_id']
    tenpo_id = info['tenpo_id']
    user_id = info['user_id']

    if not tenant_id or not user_id:
        flash('セッション情報が不正です', 'error')
        return redirect(url_for('auth.select_login'))

    if 'file' not in request.files:
        flash('ファイルが選択されていません', 'error')
        return redirect(request.url)

    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        flash('ファイルが選択されていないか、許可されていない形式です（PNG/JPG/GIF/PDF）', 'error')
        return redirect(request.url)

    try:
        upload_dir = os.path.join('uploads', 'bank', str(tenant_id))
        filepath = save_uploaded_file(file, upload_dir)

        api_keys = get_api_keys(tenant_id, tenpo_id)
        openai_api_key = api_keys.get('openai_api_key')
        google_vision_api_key = api_keys.get('google_vision_api_key')
        azure_adi_endpoint = api_keys.get('azure_document_intelligence_endpoint')
        azure_adi_key = api_keys.get('azure_document_intelligence_key')
        has_any_key = openai_api_key or google_vision_api_key or (azure_adi_endpoint and azure_adi_key)

        db = SessionLocal()
        try:
            stmt = TBankStatement(
                tenant_id=tenant_id,
                tenpo_id=tenpo_id,
                uploaded_by=user_id,
                画像パス=filepath,
                ステータス='pending' if not has_any_key else 'processing',
            )
            db.add(stmt)
            db.commit()
            stmt_id = stmt.id
        finally:
            db.close()

        # 列定義を収集
        column_def = None
        used_template_id = None  # 使用したテンプレートID（学習データの分離管理用）

        # 既存テンプレートを読み込んだ場合
        load_template_id = request.form.get('load_template_id', '')
        if load_template_id:
            try:
                used_template_id = int(load_template_id)
            except (ValueError, TypeError):
                pass

        col1_role = request.form.get('col1_role', '')
        if col1_role:  # 列定義が送信された場合
            column_def = {
                'columns': [
                    {
                        'index': i,
                        'name': request.form.get(f'col{i}_name', ''),
                        'type': request.form.get(f'col{i}_role', 'ignore'),  # ocr.py側は'type'を参照
                        'note': request.form.get(f'col{i}_note', '')
                    }
                    for i in range(1, 7)
                ],
                'note': request.form.get('col_note', '')
            }
            # テンプレート保存
            if request.form.get('save_template') and request.form.get('template_name'):
                db2 = SessionLocal()
                try:
                    if request.form.get('set_default'):
                        db2.query(TBankColumnTemplate).filter(
                            TBankColumnTemplate.tenant_id == tenant_id
                        ).update({'is_default': False})
                    tmpl = TBankColumnTemplate(
                        tenant_id=tenant_id,
                        テンプレート名=request.form.get('template_name'),
                        列1_役割=request.form.get('col1_role'), 列1_名称=request.form.get('col1_name'),
                        列2_役割=request.form.get('col2_role'), 列2_名称=request.form.get('col2_name'),
                        列3_役割=request.form.get('col3_role'), 列3_名称=request.form.get('col3_name'),
                        列4_役割=request.form.get('col4_role'), 列4_名称=request.form.get('col4_name'),
                        列5_役割=request.form.get('col5_role'), 列5_名称=request.form.get('col5_name'),
                        列6_役割=request.form.get('col6_role'), 列6_名称=request.form.get('col6_name'),
                        補足指示=request.form.get('col_note'),
                        is_default=bool(request.form.get('set_default'))
                    )
                    db2.add(tmpl)
                    db2.commit()
                    used_template_id = tmpl.id  # 新規保存したテンプレートのIDを使用
                except Exception:
                    pass
                finally:
                    db2.close()

        # TBankStatementにtemplate_idを保存
        if used_template_id:
            db3 = SessionLocal()
            try:
                s = db3.query(TBankStatement).filter(TBankStatement.id == stmt_id).first()
                if s:
                    s.template_id = used_template_id
                    db3.commit()
            except Exception:
                pass
            finally:
                db3.close()

        if not has_any_key:
            flash('APIキーが未設定のため、OCR処理をスキップしました。設定画面からOpenAIまたはGoogle Cloud Vision APIキーを設定してください。', 'warning')
        else:
            t = threading.Thread(
                target=_run_ocr_background,
                args=(stmt_id, filepath, openai_api_key, tenant_id),
                kwargs={'google_vision_api_key': google_vision_api_key, 'column_def': column_def,
                        'tenpo_id': tenpo_id, 'template_id': used_template_id,
                        'azure_document_intelligence_endpoint': azure_adi_endpoint,
                        'azure_document_intelligence_key': azure_adi_key},
                daemon=True
            )
            t.start()
            if azure_adi_endpoint and azure_adi_key:
                key_info = 'Azure Document Intelligence + GPT-4o'
            elif google_vision_api_key and openai_api_key:
                key_info = 'Google Cloud Vision API + GPT-4o'
            elif google_vision_api_key:
                key_info = 'Google Cloud Vision API'
            else:
                key_info = 'GPT-4o'
            flash(f'通帳をアップロードしました。OCR処理（{key_info}）をバックグラウンドで実行中です...', 'success')

        return redirect(url_for('voucher_bank.detail', stmt_id=stmt_id))

    except Exception as e:
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(request.url)


# ============================================================
# OCRステータス確認API（詳細画面の自動リロード用）
# ============================================================
@bp.route('/<int:stmt_id>/status')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def status(stmt_id):
    info = get_session_info()
    tenant_id = info['tenant_id']
    db = SessionLocal()
    try:
        stmt = db.query(TBankStatement).filter(
            TBankStatement.id == stmt_id,
            TBankStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            return jsonify({'status': 'error'})
        return jsonify({'status': stmt.ステータス or 'pending'})
    finally:
        db.close()


# ============================================================
# 通帳詳細
# ============================================================
@bp.route('/<int:stmt_id>')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def detail(stmt_id):
    info = get_session_info()
    tenant_id = info['tenant_id']

    db = SessionLocal()
    try:
        stmt = db.query(TBankStatement).filter(
            TBankStatement.id == stmt_id,
            TBankStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            flash('明細が見つかりません', 'error')
            return redirect(url_for('voucher_bank.index'))
        transactions = db.query(TBankTransaction).filter(
            TBankTransaction.statement_id == stmt_id
        ).order_by(TBankTransaction.行番号).all()
    finally:
        db.close()

    return render_template('voucher_bank_detail.html', stmt=stmt, transactions=transactions)


# ============================================================
# CSV出力
# ============================================================
@bp.route('/<int:stmt_id>/csv')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def export_csv(stmt_id):
    info = get_session_info()
    tenant_id = info['tenant_id']

    db = SessionLocal()
    try:
        stmt = db.query(TBankStatement).filter(
            TBankStatement.id == stmt_id,
            TBankStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            flash('明細が見つかりません', 'error')
            return redirect(url_for('voucher_bank.index'))
        transactions = db.query(TBankTransaction).filter(
            TBankTransaction.statement_id == stmt_id
        ).order_by(TBankTransaction.行番号).all()
    finally:
        db.close()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(['銀行名', stmt.銀行名 or ''])
    writer.writerow(['支店名', stmt.支店名 or ''])
    writer.writerow(['口座種別', stmt.口座種別 or ''])
    writer.writerow(['口座番号', stmt.口座番号 or ''])
    writer.writerow(['口座名義', stmt.口座名義 or ''])
    writer.writerow(['期間', f"{stmt.期間_開始 or ''} ～ {stmt.期間_終了 or ''}"])
    writer.writerow([])
    writer.writerow(['日付', '摘要', '入金', '出金', '残高', '備考'])

    for t in transactions:
        parts = [p for p in [t.摘要, t.手書き摘要] if p]
        description = ' / '.join(parts)
        writer.writerow([
            t.日付 or '',
            description,
            int(t.入金) if t.入金 is not None else '',
            int(t.出金) if t.出金 is not None else '',
            int(t.残高) if t.残高 is not None else '',
            t.備考 or '',
        ])

    output.seek(0)
    bom = '\ufeff'
    csv_bytes = io.BytesIO((bom + output.getvalue()).encode('utf-8'))
    filename = f"通帳明細_{stmt.銀行名 or 'bank'}_{stmt_id}.csv"
    return send_file(csv_bytes, mimetype='text/csv; charset=utf-8',
                     as_attachment=True, download_name=filename)


# ============================================================
# Excel出力
# ============================================================
@bp.route('/<int:stmt_id>/excel')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def export_excel(stmt_id):
    info = get_session_info()
    tenant_id = info['tenant_id']

    db = SessionLocal()
    try:
        stmt = db.query(TBankStatement).filter(
            TBankStatement.id == stmt_id,
            TBankStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            flash('明細が見つかりません', 'error')
            return redirect(url_for('voucher_bank.index'))
        transactions = db.query(TBankTransaction).filter(
            TBankTransaction.statement_id == stmt_id
        ).order_by(TBankTransaction.行番号).all()
    finally:
        db.close()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxlがインストールされていません。CSVをご利用ください。', 'error')
        return redirect(url_for('voucher_bank.detail', stmt_id=stmt_id))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '通帳明細'

    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    info_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    info_rows = [
        ('銀行名', stmt.銀行名 or ''),
        ('支店名', stmt.支店名 or ''),
        ('口座種別', stmt.口座種別 or ''),
        ('口座番号', stmt.口座番号 or ''),
        ('口座名義', stmt.口座名義 or ''),
        ('期間', f"{stmt.期間_開始 or ''} ～ {stmt.期間_終了 or ''}"),
    ]
    for r, (label, value) in enumerate(info_rows, 1):
        ws.cell(r, 1, label).fill = info_fill
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value)

    header_row = len(info_rows) + 2
    headers = ['日付', '摘要', '入金', '出金', '残高', '備考']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(header_row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, t in enumerate(transactions):
        row = header_row + 1 + i
        parts = [p for p in [t.摘要, t.手書き摘要] if p]
        description = ' / '.join(parts)
        values = [
            t.日付 or '',
            description,
            int(t.入金) if t.入金 is not None else '',
            int(t.出金) if t.出金 is not None else '',
            int(t.残高) if t.残高 is not None else '',
            t.備考 or '',
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row, c, v)
            cell.border = border
            if c in (3, 4, 5) and v != '':
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"通帳明細_{stmt.銀行名 or 'bank'}_{stmt_id}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ============================================================
# 手動修正 API（明細行の一括更新）
# ============================================================
@bp.route('/<int:stmt_id>/update_transactions', methods=['POST'])
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def update_transactions(stmt_id):
    """明細テーブルの手動修正を受け取りDBを更新する"""
    info = get_session_info()
    tenant_id = info['tenant_id']

    db = SessionLocal()
    try:
        stmt = db.query(TBankStatement).filter(
            TBankStatement.id == stmt_id,
            TBankStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            return jsonify({'success': False, 'error': '明細が見つかりません'}), 404

        data = request.get_json()
        if not data or 'transactions' not in data:
            return jsonify({'success': False, 'error': 'データが不正です'}), 400

        # 修正前の摘要を保存（学習用）
        # クライアントから「元の摘要（original_description）」を送信してもらう
        # 送信されない場合は現在DBの摘要を使用
        existing_rows = db.query(TBankTransaction).filter(
            TBankTransaction.statement_id == stmt_id
        ).order_by(TBankTransaction.行番号).all()
        existing_desc_map = {row.行番号: row.摘要 for row in existing_rows}

        # 既存の明細を全削除して再登録
        db.query(TBankTransaction).filter(TBankTransaction.statement_id == stmt_id).delete()

        def to_float_or_none(v):
            if v is None or str(v).strip() == '':
                return None
            try:
                return float(str(v).replace(',', '').replace('¥', '').replace('円', '').strip())
            except (ValueError, TypeError):
                return None

        learning_saved = 0
        for i, t in enumerate(data['transactions'], 1):
            new_desc = t.get('description') or None

            # 学習データの記録：元の摘要と修正後の摘要を比較
            # クライアントからoriginal_descriptionが送信された場合はそれを使用
            original_desc = t.get('original_description') or existing_desc_map.get(i)
            if original_desc and new_desc and original_desc != new_desc:
                try:
                    _save_learning_data(db, tenant_id, original_desc, new_desc)
                    learning_saved += 1
                except Exception as le:
                    print(f'[学習] 記録エラー: {le}')

            row = TBankTransaction(
                statement_id=stmt_id,
                tenant_id=tenant_id,
                日付=t.get('date') or None,
                摘要=new_desc,
                手書き摘要=t.get('handwritten_description') or None,
                入金=to_float_or_none(t.get('deposit')),
                出金=to_float_or_none(t.get('withdrawal')),
                残高=to_float_or_none(t.get('balance')),
                備考=t.get('note') or None,
                行番号=i,
            )
            db.add(row)

        db.commit()
        msg = f'保存しました（{len(data["transactions"])}行'
        if learning_saved > 0:
            msg += f'、{learning_saved}件の摘要修正を学習しました'
        msg += '）'
        return jsonify({'success': True, 'count': len(data['transactions']), 'learning_saved': learning_saved, 'message': msg})

    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        db.close()
