# -*- coding: utf-8 -*-
"""
証憑データ化アプリ - クレジット明細モード Blueprint
クレジット明細画像のアップロード、OCR処理（バックグラウンド）、一覧表示、CSV/Excel出力
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import os
import io
import csv
import threading
from datetime import datetime
from app.db import SessionLocal
from app.models_voucher import TCreditStatement, TCreditTransaction
from app.models_login import TTenpo, TTenant, TKanrisha
from app.utils.decorators import require_roles, ROLES
from app.utils.voucher.ocr import process_credit_statement_image, save_uploaded_file

bp = Blueprint('voucher_credit', __name__, url_prefix='/voucher/credit')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_session_info():
    tenpo_id = session.get('store_id') or session.get('tenpo_id')
    return {
        'tenant_id': session.get('tenant_id'),
        'tenpo_id': tenpo_id,
        'user_id': session.get('user_id'),
        'role': session.get('role'),
    }


def get_openai_api_key(tenant_id, tenpo_id):
    """店舗 → テナント → システム管理者の順でOpenAI APIキーを取得"""
    keys = get_api_keys(tenant_id, tenpo_id)
    return keys.get('openai_api_key')


def get_api_keys(tenant_id, tenpo_id):
    """店舗 → テナント → システム管理者の順で各種 APIキーを取得"""
    db = SessionLocal()
    result = {'openai_api_key': None, 'google_vision_api_key': None, 'google_api_key': None, 'anthropic_api_key': None}
    try:
        if tenpo_id:
            tenpo = db.query(TTenpo).filter(TTenpo.id == tenpo_id).first()
            if tenpo:
                if getattr(tenpo, 'openai_api_key', None):
                    result['openai_api_key'] = tenpo.openai_api_key
                if getattr(tenpo, 'google_vision_api_key', None):
                    result['google_vision_api_key'] = tenpo.google_vision_api_key
                if getattr(tenpo, 'google_api_key', None):
                    result['google_api_key'] = tenpo.google_api_key
                if getattr(tenpo, 'anthropic_api_key', None):
                    result['anthropic_api_key'] = tenpo.anthropic_api_key
        if tenant_id:
            tenant = db.query(TTenant).filter(TTenant.id == tenant_id).first()
            if tenant:
                if not result['openai_api_key'] and getattr(tenant, 'openai_api_key', None):
                    result['openai_api_key'] = tenant.openai_api_key
                if not result['google_vision_api_key'] and getattr(tenant, 'google_vision_api_key', None):
                    result['google_vision_api_key'] = tenant.google_vision_api_key
                if not result['google_api_key'] and getattr(tenant, 'google_api_key', None):
                    result['google_api_key'] = tenant.google_api_key
                if not result['anthropic_api_key'] and getattr(tenant, 'anthropic_api_key', None):
                    result['anthropic_api_key'] = tenant.anthropic_api_key
        # システム管理者から未設定のAPIキーをフォールバック
        needs_fallback = not all([result['openai_api_key'], result['google_vision_api_key'], result['google_api_key'], result['anthropic_api_key']])
        if needs_fallback:
            sys_admins = db.query(TKanrisha).filter(
                TKanrisha.role == 'system_admin'
            ).all()
            for sys_admin in sys_admins:
                if not result['openai_api_key'] and getattr(sys_admin, 'openai_api_key', None):
                    result['openai_api_key'] = sys_admin.openai_api_key
                if not result['google_vision_api_key'] and getattr(sys_admin, 'google_vision_api_key', None):
                    result['google_vision_api_key'] = sys_admin.google_vision_api_key
                if not result['google_api_key'] and getattr(sys_admin, 'google_api_key', None):
                    result['google_api_key'] = sys_admin.google_api_key
                if not result['anthropic_api_key'] and getattr(sys_admin, 'anthropic_api_key', None):
                    result['anthropic_api_key'] = sys_admin.anthropic_api_key
                if all([result['openai_api_key'], result['google_vision_api_key'], result['google_api_key'], result['anthropic_api_key']]):
                    break
    except Exception:
        pass
    finally:
        db.close()
    return result


def _run_ocr_background(stmt_id, filepath, api_key, tenant_id, google_vision_api_key=None):
    """バックグラウンドスレッドでOCR処理を実行してDBを更新する"""
    db = SessionLocal()
    try:
        # ステータスを「処理中」に更新
        stmt = db.query(TCreditStatement).filter(TCreditStatement.id == stmt_id).first()
        if not stmt:
            return
        stmt.ステータス = 'processing'
        db.commit()

        # OCR実行（時間がかかる処理）
        ocr_result = process_credit_statement_image(filepath, api_key=api_key, google_vision_api_key=google_vision_api_key)

        # 結果をDBに保存
        stmt = db.query(TCreditStatement).filter(TCreditStatement.id == stmt_id).first()
        if not stmt:
            return

        stmt.OCR結果_生データ = ocr_result.get('raw_text', '')
        stmt.カード会社名 = ocr_result.get('card_company')
        stmt.カード名 = ocr_result.get('card_name')
        stmt.会員名 = ocr_result.get('member_name')
        stmt.明細年月 = ocr_result.get('statement_month')
        stmt.支払日 = ocr_result.get('payment_date')
        stmt.利用総額 = ocr_result.get('total_amount')
        stmt.ステータス = 'completed'
        db.flush()

        # 既存の明細を削除して再登録
        db.query(TCreditTransaction).filter(TCreditTransaction.statement_id == stmt_id).delete()

        for i, t in enumerate(ocr_result.get('transactions', []), 1):
            row = TCreditTransaction(
                statement_id=stmt_id,
                tenant_id=tenant_id,
                利用日=t.get('date'),
                利用店名=t.get('store_name'),
                利用者=t.get('user_name'),
                利用金額=t.get('amount'),
                分割回数=t.get('installment'),
                備考=t.get('note'),
                行番号=i,
            )
            db.add(row)

        db.commit()

    except Exception as e:
        try:
            stmt = db.query(TCreditStatement).filter(TCreditStatement.id == stmt_id).first()
            if stmt:
                stmt.ステータス = 'error'
                stmt.OCR結果_生データ = f'OCRエラー: {str(e)}'
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


# ============================================================
# クレジット明細一覧
# ============================================================
@bp.route('/')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def index():
    info = get_session_info()
    tenant_id = info['tenant_id']
    tenpo_id = info['tenpo_id']
    role = info['role']

    if not tenant_id:
        flash('テナントが選択されていません', 'error')
        return redirect(url_for('auth.select_login'))

    db = SessionLocal()
    try:
        q = db.query(TCreditStatement).filter(TCreditStatement.tenant_id == tenant_id)
        if role in (ROLES['ADMIN'], ROLES['EMPLOYEE']) and tenpo_id:
            q = q.filter(TCreditStatement.tenpo_id == tenpo_id)
        statements = q.order_by(TCreditStatement.created_at.desc()).all()
    finally:
        db.close()

    return render_template('voucher_credit_list.html', statements=statements)


# ============================================================
# クレジット明細アップロード（即時リダイレクト、OCRはバックグラウンド）
# ============================================================
@bp.route('/upload', methods=['GET', 'POST'])
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def upload():
    if request.method == 'GET':
        return render_template('voucher_credit_upload.html')

    info = get_session_info()
    tenant_id = info['tenant_id']
    tenpo_id = info['tenpo_id']
    user_id = info['user_id']

    if not tenant_id or not user_id:
        flash('セッション情報が不正です', 'error')
        return redirect(url_for('auth.select_login'))

    if 'file' not in request.files:
        flash('ファイルが選択されていません', 'error')
        return redirect(request.url)

    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        flash('ファイルが選択されていないか、許可されていない形式です（PNG/JPG/GIF/PDF）', 'error')
        return redirect(request.url)

    try:
        upload_dir = os.path.join('uploads', 'credit', str(tenant_id))
        filepath = save_uploaded_file(file, upload_dir)

        api_keys = get_api_keys(tenant_id, tenpo_id)
        openai_api_key = api_keys.get('openai_api_key')
        google_vision_api_key = api_keys.get('google_vision_api_key')
        has_any_key = openai_api_key or google_vision_api_key

        # DBに「待機中」として即座に保存
        db = SessionLocal()
        try:
            stmt = TCreditStatement(
                tenant_id=tenant_id,
                tenpo_id=tenpo_id,
                uploaded_by=user_id,
                画像パス=filepath,
                ステータス='pending' if not has_any_key else 'processing',
            )
            db.add(stmt)
            db.commit()
            stmt_id = stmt.id
        finally:
            db.close()

        if not has_any_key:
            flash('APIキーが未設定のため、OCR処理をスキップしました。設定画面からOpenAIまたはGoogle Cloud Vision APIキーを設定してください。', 'warning')
        else:
            # バックグラウンドスレッドでOCR処理を開始
            t = threading.Thread(
                target=_run_ocr_background,
                args=(stmt_id, filepath, openai_api_key, tenant_id),
                kwargs={'google_vision_api_key': google_vision_api_key},
                daemon=True
            )
            t.start()
            key_info = 'Google Cloud Vision API + GPT-4o' if google_vision_api_key and openai_api_key else ('Google Cloud Vision API' if google_vision_api_key else 'GPT-4o')
            flash(f'クレジット明細をアップロードしました。OCR処理（{key_info}）をバックグラウンドで実行中です...', 'success')

        return redirect(url_for('voucher_credit.detail', stmt_id=stmt_id))

    except Exception as e:
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(request.url)


# ============================================================
# OCRステータス確認API（詳細画面の自動リロード用）
# ============================================================
@bp.route('/<int:stmt_id>/status')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def status(stmt_id):
    info = get_session_info()
    tenant_id = info['tenant_id']
    db = SessionLocal()
    try:
        stmt = db.query(TCreditStatement).filter(
            TCreditStatement.id == stmt_id,
            TCreditStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            return jsonify({'status': 'error'})
        return jsonify({'status': stmt.ステータス or 'pending'})
    finally:
        db.close()


# ============================================================
# クレジット明細詳細
# ============================================================
@bp.route('/<int:stmt_id>')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def detail(stmt_id):
    info = get_session_info()
    tenant_id = info['tenant_id']

    db = SessionLocal()
    try:
        stmt = db.query(TCreditStatement).filter(
            TCreditStatement.id == stmt_id,
            TCreditStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            flash('明細が見つかりません', 'error')
            return redirect(url_for('voucher_credit.index'))
        transactions = db.query(TCreditTransaction).filter(
            TCreditTransaction.statement_id == stmt_id
        ).order_by(TCreditTransaction.行番号).all()
    finally:
        db.close()

    return render_template('voucher_credit_detail.html', stmt=stmt, transactions=transactions)


# ============================================================
# CSV出力
# ============================================================
@bp.route('/<int:stmt_id>/csv')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def export_csv(stmt_id):
    info = get_session_info()
    tenant_id = info['tenant_id']

    db = SessionLocal()
    try:
        stmt = db.query(TCreditStatement).filter(
            TCreditStatement.id == stmt_id,
            TCreditStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            flash('明細が見つかりません', 'error')
            return redirect(url_for('voucher_credit.index'))
        transactions = db.query(TCreditTransaction).filter(
            TCreditTransaction.statement_id == stmt_id
        ).order_by(TCreditTransaction.行番号).all()
    finally:
        db.close()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(['カード会社名', stmt.カード会社名 or ''])
    writer.writerow(['カード名', stmt.カード名 or ''])
    writer.writerow(['会員名', stmt.会員名 or ''])
    writer.writerow(['明細年月', stmt.明細年月 or ''])
    writer.writerow(['支払日', stmt.支払日 or ''])
    writer.writerow(['利用総額', int(stmt.利用総額) if stmt.利用総額 is not None else ''])
    writer.writerow([])
    writer.writerow(['利用日', '利用店名', '利用者', '利用金額', '分割回数', '備考'])

    for t in transactions:
        writer.writerow([
            t.利用日 or '',
            t.利用店名 or '',
            t.利用者 or '',
            int(t.利用金額) if t.利用金額 is not None else '',
            t.分割回数 or '',
            t.備考 or '',
        ])

    output.seek(0)
    bom = '\ufeff'
    csv_bytes = io.BytesIO((bom + output.getvalue()).encode('utf-8'))
    filename = f"クレジット明細_{stmt.カード会社名 or 'credit'}_{stmt_id}.csv"
    return send_file(csv_bytes, mimetype='text/csv; charset=utf-8',
                     as_attachment=True, download_name=filename)


# ============================================================
# Excel出力
# ============================================================
@bp.route('/<int:stmt_id>/excel')
@require_roles(ROLES['SYSTEM_ADMIN'], ROLES['TENANT_ADMIN'], ROLES['ADMIN'], ROLES['EMPLOYEE'])
def export_excel(stmt_id):
    info = get_session_info()
    tenant_id = info['tenant_id']

    db = SessionLocal()
    try:
        stmt = db.query(TCreditStatement).filter(
            TCreditStatement.id == stmt_id,
            TCreditStatement.tenant_id == tenant_id
        ).first()
        if not stmt:
            flash('明細が見つかりません', 'error')
            return redirect(url_for('voucher_credit.index'))
        transactions = db.query(TCreditTransaction).filter(
            TCreditTransaction.statement_id == stmt_id
        ).order_by(TCreditTransaction.行番号).all()
    finally:
        db.close()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxlがインストールされていません。CSVをご利用ください。', 'error')
        return redirect(url_for('voucher_credit.detail', stmt_id=stmt_id))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'クレジット明細'

    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    info_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    info_rows = [
        ('カード会社名', stmt.カード会社名 or ''),
        ('カード名', stmt.カード名 or ''),
        ('会員名', stmt.会員名 or ''),
        ('明細年月', stmt.明細年月 or ''),
        ('支払日', stmt.支払日 or ''),
        ('利用総額', int(stmt.利用総額) if stmt.利用総額 is not None else ''),
    ]
    for r, (label, value) in enumerate(info_rows, 1):
        ws.cell(r, 1, label).fill = info_fill
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value)

    header_row = len(info_rows) + 2
    headers = ['利用日', '利用店名', '利用者', '利用金額', '分割回数', '備考']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(header_row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, t in enumerate(transactions):
        row = header_row + 1 + i
        values = [
            t.利用日 or '',
            t.利用店名 or '',
            t.利用者 or '',
            int(t.利用金額) if t.利用金額 is not None else '',
            t.分割回数 or '',
            t.備考 or '',
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row, c, v)
            cell.border = border
            if c == 4 and v != '':
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"クレジット明細_{stmt.カード会社名 or 'credit'}_{stmt_id}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
